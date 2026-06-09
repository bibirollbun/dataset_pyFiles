# KẾT NỐI VÀ XỬ LÝ DỮ LIỆU 
import pyodbc  # Kết nối cơ sở dữ liệu (SQL Server, v.v.)
import pandas as pd  # Xử lý dữ liệu dạng bảng (DataFrame)

# === BIỂU ĐỒ VÀ ĐỒ HỌA ===
import matplotlib.pyplot as plt  # Vẽ biểu đồ (line, bar, histogram,...)
import matplotlib.cm as cm  # Dùng để lấy colormap (phối màu)
import matplotlib.colors as mcolors  # Quản lý màu sắc (RGB, HEX,...)
import matplotlib.patches as mpatches  # Tạo mẫu màu dùng cho chú thích (legend)

# === HỆ THỐNG VÀ TOÁN HỌC ===
import sys  # Truy cập thông tin hệ thống (ví dụ: sys.path)
import numpy as np  # Thư viện toán học, xử lý mảng và thống kê

# === XUẤT FILE EXCEL ===
import xlsxwriter  # Ghi dữ liệu và biểu đồ ra file Excel (.xlsx)

# === TỔ HỢP VÀ KẾT HỢP BIẾN ===
from itertools import combinations  # Sinh các tổ hợp biến để so sánh/cross feature

# === LƯU ẢNH HOẶC FILE VÀO BỘ NHỚ TẠM ===
from io import BytesIO  # Dùng để lưu ảnh Excel/chart vào RAM thay vì file

# === MÔ HÌNH THỐNG KÊ ===
import statsmodels.api as sm  # Hồi quy logistic, OLS, thống kê mô hình



# FUNCTION: BINNING BIẾN
def binning_all_variables(df, threshold=0.05, custom_bins=None, feature=None):
    df = df.copy()  # Sao chép DataFrame gốc để tránh thay đổi ngoài ý muốn

    columns_to_process = [feature] if feature else df.columns  # Nếu có truyền feature thì chỉ xử lý biến đó, nếu không thì xử lý toàn bộ

    for column in columns_to_process:
        if column in ['case_id','date_decision', 'WEEK_NUM', 'MONTH', 'target']:
            continue  # Bỏ qua các biến đặc biệt không cần binning

        missing_mask = df[column].isna()  # Xác định vị trí các giá trị missing
        add_missing = False  # Cờ để ghi nhớ xem có missing không (để thêm vào danh sách nhóm sau)

        # --- Xử lý biến phân loại ---
        if df[column].dtype == object or df[column].dtype.name == 'category':
            grp_col = 'GRP_' + column  # Tên cột mới sau khi gộp nhóm
            df[grp_col] = df[column].astype(object)  # Copy giá trị gốc sang cột nhóm
            if missing_mask.any():
                df.loc[missing_mask, grp_col] = 'Missing'  # Gán label 'Missing' cho các giá trị thiếu
                add_missing = True

        else:
            # --- Xử lý biến số ---
            if custom_bins and column in custom_bins:
                bins = [-np.inf] + custom_bins[column] + [np.inf]  # Nếu có custom bins thì dùng
            else:
                value_counts = df[column].value_counts(normalize=True)  # Tính phân bố tỷ lệ giá trị
                sorted_vals = np.sort(df[column].dropna().unique())  # Sắp xếp giá trị không missing
                cumulative = 0  # Tích lũy để gộp nhóm nhỏ lại
                bins = [-np.inf]  # Khởi tạo mốc dưới đầu tiên
                for v in sorted_vals:
                    prop = value_counts.get(v, 0)  # Lấy tỷ lệ xuất hiện của v
                    if prop > threshold:
                        bins.append(v)  # Nếu đủ lớn thì đưa vào nhóm mới
                    else:
                        cumulative += prop  # Gộp vào nhóm hiện tại nếu nhỏ
                        if cumulative >= threshold:
                            bins.append(v)
                            cumulative = 0
                bins.append(np.inf)  # Mốc trên cùng
                bins = sorted(set(bins))  # Loại bỏ trùng và sắp xếp

        
            grp_col = 'GRP_' + column
            df[grp_col] = pd.cut(df[column], bins=bins, include_lowest=True, duplicates='drop')  # Phân nhóm
            if missing_mask.any():
                df[grp_col] = df[grp_col].astype(object)  # Ép kiểu để gán 'Missing'
                df.loc[missing_mask, grp_col] = 'Missing'
                add_missing = True

        # --- Sắp xếp lại các nhóm theo thứ tự hợp lý ---
        def extract_lb(x): return x.left if isinstance(x, pd.Interval) else np.inf  # Trích mốc dưới nếu là khoảng

        if df[column].dtype == object or df[column].dtype.name == 'category':
            categories = [x for x in df['GRP_' + column].unique() if x != 'Missing']  # Nhóm không missing
        else:
            ivals = [c for c in df['GRP_' + column].unique() if isinstance(c, pd.Interval)]
            categories = sorted(ivals, key=extract_lb)  # Sắp theo mốc dưới

        if add_missing:
            categories.append('Missing')  # Đưa 'Missing' vào cuối danh sách nhóm

        df['GRP_' + column] = pd.Categorical(df['GRP_' + column], categories=categories, ordered=True)  # Gán lại với nhóm đã sắp xếp

    return df  # Trả về DataFrame sau khi binning



# VÍ DỤ TẠO DATAFRAME ỨNG DỤNG CÁC FUNCTION (kích thước data phân tích lớn sử dụng trình bày sẽ khó tối ưu)

np.random.seed(42)
n = 10000

# 1. case_id duy nhất
case_id = np.arange(n)

# 2. days90_310L ngẫu nhiên từ 0–10, thêm 20% NaN
days_values = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, np.nan]
prob = np.array([0.34, 0.18, 0.12, 0.06, 0.03, 0.03, 0.06, 0.05, 0.04, 0.04, 0.05, 0.20])

# Chuẩn hóa xác suất để tổng đúng 1.0
prob = prob / prob.sum()

days90_310L = np.random.choice(days_values, size=n, p=prob)

# 3. sex_738L: F 61%, M 39%
sex_738L = np.random.choice(['F', 'M'], size=n, p=[0.61, 0.39])

# 4. target: 0 chiếm 97%, 1 chiếm 3%
target = np.random.choice([0, 1], size=n, p=[0.97, 0.03])

# 5. Gộp lại
df = pd.DataFrame({
    'case_id': case_id,
    'days90_310L': days90_310L,
    'sex_738L': sex_738L,
    'target': target
})

df.head()



# SỬ DỤNG FUNCTION BINNING VỚI BIẾN days90_310L
df = binning_all_variables(df, threshold=0.05, custom_bins=None, feature='days90_310L')
df


# FUNCTION TÍNH WOE VÀ IV
def caculate_WOE_IV(df, feature, target):
    # Nhóm theo giá trị của biến 'feature' và tính tổng số quan sát và tổng số event (target = 1)
    df = df.groupby(feature, observed=False)[target].agg(['count', 'sum']).reset_index()
    
    # Đặt lại tên cột: bin (giá trị của feature, binning), num_of_obs (số quan sát của bin), num_of_event (số event = 1, tính tổng)
    df.columns = ['bin', 'num_of_obs', 'num_of_event']
    
    # Tính số non-event (target = 0)
    df['num_of_non_event'] = df['num_of_obs'] - df['num_of_event']
    
    # Ép kiểu về float để tránh lỗi khi chia
    df['num_of_event'] = df['num_of_event'].astype(float)
    df['num_of_non_event'] = df['num_of_non_event'].astype(float)

    # Tạo 2 cột sao lưu dùng để điều chỉnh nếu có event hoặc non-event bằng 0
    df['num_of_non_event_c'] = df['num_of_non_event'].copy()
    df['num_of_event_c'] = df['num_of_event'].copy()

    # Tổng số event và non-event toàn bộ tập
    total_non_event = df['num_of_non_event'].sum()
    total_event = df['num_of_event'].sum()

    # Nếu bất kỳ bin nào có số event hoặc non-event = 0 → cộng 0.5 để tránh chia cho 0 hoặc log(0)
    mask = (df['num_of_event'] == 0) | (df['num_of_non_event'] == 0)
    df.loc[mask, 'num_of_event_c'] += 0.5
    df.loc[mask, 'num_of_non_event_c'] += 0.5

    # Tính tỉ lệ non-event và event trong mỗi bin
    df['prct_non_event'] = df['num_of_non_event_c'] / total_non_event
    df['prct_event'] = df['num_of_event_c'] / total_event

    # Tính WOE: log(tỉ lệ non-event / tỉ lệ event)
    df['WOE'] = np.log(df['prct_non_event'] / df['prct_event'])

    # Tính IV: (tỉ lệ non-event - tỉ lệ event) * WOE
    df['IV'] = (df['prct_non_event'] - df['prct_event']) * df['WOE']
    
    # Tổng IV của biến → dùng để đo mức độ phân biệt của biến
    IV = df['IV'].sum()

    # Đổi bin về dạng chuỗi để dễ biểu diễn hoặc vẽ biểu đồ
    df['bin'] = df['bin'].astype(str)

    # Xoá các cột trung gian đã dùng để tính toán
    df = df.drop(columns=['num_of_non_event_c', 'num_of_event_c'])

    # Tính tỷ lệ event trong mỗi bin
    df['event_rate'] = df['num_of_event'] / df['num_of_obs']

    # Tính tỷ lệ quan sát trong mỗi bin (dạng phần trăm)
    df['prct_obs'] = 100 * df['num_of_obs'] / df['num_of_obs'].sum()

    # Trả về dataframe kết quả và giá trị IV
    return df, IV



# SỬ DỤNG FUNCTION TÍNH WOE VÀ IV VỚI days90_310L
df_woe = caculate_WOE_IV(df, 'GRP_days90_310L', 'target')
df_woe


#FUNCTION TẠO BẢNG WOE 
def woe_feature_table(data, custom_bins, feature):
    # Binning dữ liệu theo biến 'feature' với ngưỡng tối thiểu cho mỗi bin là 5% và sử dụng custom_bins
    df_binned = binning_all_variables(data, threshold=0.05, custom_bins=custom_bins, feature=feature)

    # Tính bảng WOE và giá trị IV với biến đã bin là 'GRP_<feature>' và biến mục tiêu là 'target'
    WOE_TABLE, IV = caculate_WOE_IV(df_binned, f'GRP_{feature}', 'target')

    # Chọn các cột cần thiết để trình bày bảng WOE
    df = WOE_TABLE[['bin', 'num_of_obs', 'prct_obs', 'num_of_non_event', 'num_of_event', 'event_rate', 'WOE', 'IV']]

    # Đặt lại tên cột cho dễ hiểu và phù hợp với báo cáo
    df.columns = ['bin', 'count', 'count(%)', 'non_event', 'event', 'event_rate', 'WOE', 'IV']

    # Tạo dòng tổng cộng để tổng hợp tất cả các bin
    total_row = pd.DataFrame({
        'bin': [''],
        'count': [df['count'].sum()],
        'count(%)': [100],
        'non_event': [df['non_event'].sum()],
        'event': [df['event'].sum()],
        'event_rate': [df['event'].sum() / df['count'].sum()],
        'WOE': [''],  # Không tính WOE tổng
        'IV': [df['IV'].sum()]  # Tổng IV
    })

    # Gộp bảng chính và dòng tổng lại
    df = pd.concat([df, total_row], ignore_index=True)

    # Đổi index thành chuỗi để dễ hiển thị, đặt tên dòng cuối là 'TOTAL'
    df.index = df.index.astype(str)
    df.index.values[-1] = "TOTAL"

    return df


# SỬ DỤNG FUNCTION VỚI BIẾN days90_310L
df_woe_days90_310L = woe_feature_table(df, custom_bins = None, feature = 'days90_310L')
df_woe_days90_310L


# FUNCTION CROSS 2 BIẾN
# Hàm tạo biến kết hợp giữa 2 biến đầu vào và tính bảng WOE + pivot table

def cross_variables(df, feature1, feature2, custom_bins=None):
    # Thực hiện binning (chia nhóm) cho cả 2 biến đầu vào
    df = binning_all_variables(df, threshold=0.05, custom_bins=custom_bins, feature=feature1)
    df = binning_all_variables(df, threshold=0.05, custom_bins=custom_bins, feature=feature2)

    # Tạo tên cột group sau khi binning
    grp1 = 'GRP_' + feature1
    grp2 = 'GRP_' + feature2
    cross_feature = f"{feature1}_{feature2}_cross"

    # Tạo biến mới kết hợp giữa 2 biến đã binning (dạng chuỗi nối với '*')
    df[cross_feature] = df[grp1].astype(str) + '*' + df[grp2].astype(str)

    # Tính bảng WOE và IV cho biến kết hợp
    df_woe, _ = caculate_WOE_IV(df, cross_feature, 'target')

    # Tách lại tên bin của 2 biến ban đầu từ biến kết hợp
    df_woe[[grp1, grp2]] = df_woe['bin'].str.split('*', expand=True)

    # HÀM PARSE: Chuyển đổi chuỗi dạng '(a, b]' về kiểu pd.Interval để giữ thứ tự logic
    def parse_interval(x):
        try:
            if x == 'Missing':
                return x
            x = x.replace('(', '[')  # Đồng nhất dấu ngoặc
            bounds = x.strip('[]').split(',')
            return pd.Interval(float(bounds[0]), float(bounds[1]), closed='right')
        except:
            return x

    # Áp dụng chuẩn hóa lại kiểu dữ liệu cho cả 2 biến đã tách
    df_woe[grp1] = df_woe[grp1].apply(parse_interval)
    df_woe[grp2] = df_woe[grp2].apply(parse_interval)

    # Sau khi đã sử dụng, loại bỏ cột bin gốc khỏi df ban đầu để tránh dư thừa
    df.drop(columns=[grp1, grp2], inplace=True)

    # Tạo bảng pivot để thể hiện phân tích theo 2 chiều (2 biến)
    pivot_df = df_woe.pivot(index=grp1, columns=grp2, values=["event_rate", "num_of_obs", "prct_obs", "WOE"])

    # Đổi tên trục hàng và cột trong bảng pivot theo tên biến gốc
    pivot_df = pivot_df.rename_axis(index={f"GRP_{feature1}": feature1}, columns={f"GRP_{feature2}": feature2})

    # Sắp xếp lại các cột theo thứ tự logic của nhóm
    pivot_df = pivot_df.sort_index(axis=1, level=1)

    # Trả ra: df gốc đã xóa group, bảng WOE chi tiết, và bảng pivot theo 2 biến
    return df, df_woe, pivot_df



# SỬ DỤNG FUNCTION
df, df_woe_cross, df_pivot = cross_variables(df, 'days90_310L', 'sex_738L', custom_bins=None)
df


df_woe_cross


df_pivot


#FUNCTION VẼ BIỂU ĐỒ WOE CHO TỪNG BIẾN
# Hàm vẽ biểu đồ WOE stacked bar chart + đường line WOE cho một biến đã qua binning
import matplotlib.pyplot as plt  # Vẽ biểu đồ (line, bar, histogram,...)
import matplotlib.pyplot as plt  # Vẽ biểu đồ (line, bar, histogram,...)
import matplotlib.cm as cm  # Dùng để lấy colormap (phối màu)
import matplotlib.colors as mcolors  # Quản lý màu sắc (RGB, HEX,...)
import matplotlib.patches as mpatches  # Tạo mẫu màu dùng cho chú thích (legend)
def woe_plot(df, feature):
    # Loại bỏ dòng 'TOTAL' nếu có trong bảng dữ liệu WOE
    df = df[df.index != "TOTAL"]

    # Tạo colormap từ đỏ -> vàng -> xanh theo giá trị WOE
    cmap = plt.colormaps["RdYlGn"]
    norm = mcolors.Normalize(vmin=df['WOE'].min(), vmax=df['WOE'].max())  # Chuẩn hóa giá trị WOE để gán màu tương ứng

    # Tạo biểu đồ chính
    fig, ax1 = plt.subplots(figsize=(9, 7))
    bins = np.arange(len(df))  # Trục x là các bin thứ tự

    # Xác định màu của từng bin theo giá trị WOE
    colors = [cmap(norm(woe)) for woe in df['WOE']]

    # Vẽ stacked bar: event + non-event, dùng cùng màu (tùy theo giá trị WOE)
    ax1.bar(bins, df['event'], label='Event', color=colors, alpha=1)
    ax1.bar(bins, df['non_event'], label='Non-event', bottom=df['event'], color=colors, alpha=1)

    # Cấu hình trục y bên trái
    ax1.set_xlabel('Bin')
    ax1.set_ylabel('Count')
    ax1.set_xticks(bins)
    ax1.set_xticklabels(df['bin'], rotation=45, ha='right')

    # Tạo trục y bên phải và vẽ đường WOE
    ax2 = ax1.twinx()
    ax2.plot(bins, df['WOE'], color='black', marker='o', markersize=6, linestyle='-', linewidth=2, label='WOE')
    ax2.set_ylabel('WOE')

    # Thêm tiêu đề và lưới nhẹ
    ax1.grid(axis='y', linestyle='--', alpha=0.7)
    plt.title(f'WOE Plot for {feature}', fontsize=12, fontweight='bold')

    # Định nghĩa các chú thích màu đại diện cho rủi ro
    high_risk_patch = mpatches.Patch(facecolor='red', label='Rủi ro cao')
    medium_risk_patch = mpatches.Patch(facecolor='yellow', label='Rủi ro trung bình')
    low_risk_patch = mpatches.Patch(facecolor='green', label='Rủi ro thấp')

    # Hiển thị chú thích theo dạng hàng ngang phía dưới biểu đồ
    ax1.legend(
        handles=[high_risk_patch, medium_risk_patch, low_risk_patch],
        bbox_to_anchor=(0.9, -0.5),
        frameon=True,
        ncol=3
    )

    plt.tight_layout()
    plt.show()


# SỬ DỤNG FUNCTION
PLOT = woe_plot(df_woe_days90_310L, 'days90_310L')


def export_all_to_excel_1(df, custom_bins, features, output_file, df_TH):
    # Khởi tạo writer và workbook
    writer = pd.ExcelWriter(output_file, engine='xlsxwriter')
    workbook = writer.book
    worksheet = workbook.add_worksheet("WOE_All")  # Tạo sheet
    writer.sheets["WOE_All"] = worksheet

    current_row = 0  # Dòng bắt đầu ghi dữ liệu cho mỗi biến

    cmap = plt.colormaps["RdYlGn"]  # Colormap dùng cho biểu đồ WOE

    # Duyệt qua từng biến trong danh sách features
    for feature in tqdm(features, desc="Xuất biến"):
        metadata_start_row = current_row  # Lưu vị trí dòng đầu tiên cho metadata (cũng là nơi chèn biểu đồ)
        # Lấy dòng metadata tương ứng trong df_TH theo tên biến
        
        row = df_TH[df_TH["Variable"] == feature].squeeze()

# Tạo metadata từ df_TH
        metadata = [
           ["Tên biến", feature],
            ["Nguồn", row.get("Bảng", "")],
            ["Loại thông tin", row.get("Loại thông tin", "")],
            ["Description gốc", row.get("Description gốc", "")],
            ["Mô tả", row.get("Mô tả", "")],
            ["Unique Values", row.get("Unique Values", "")],
            ["Clean", row.get("Clean", "")],
            ["Phân tích", row.get("Phân tích", "")],
            ["trend WOE", ""],
            ["Note", row.get("Note", "")],
            ["Splits", row.get("Splits", "")]
        ]
        

        # Chuyển metadata sang DataFrame để ghi nhanh vào Excel
        metadata_df = pd.DataFrame(metadata)
        metadata_df.to_excel(writer, sheet_name="WOE_All", startrow=current_row, startcol=0, header=False, index=False)

        current_row += len(metadata)  # Di chuyển dòng xuống sau metadata

        # Tính bảng WOE cho biến hiện tại
        woe_df = woe_feature_table(df, custom_bins, feature)

        # Ghi bảng WOE vào Excel
        woe_df.to_excel(writer, sheet_name="WOE_All", startrow=current_row, startcol=0, index=False)

        # Vẽ biểu đồ WOE cho biến hiện tại
        fig = plt.figure(figsize=(6, 4))
        temp_df = woe_df[woe_df.index != "TOTAL"]  # Bỏ dòng TOTAL
        norm = mcolors.Normalize(vmin=temp_df['WOE'].min(), vmax=temp_df['WOE'].max())
        colors = [cmap(norm(woe)) for woe in temp_df['WOE']]

        bins = np.arange(len(temp_df))
        ax1 = fig.add_subplot(111)
        ax1.bar(bins, temp_df['event'], color=colors)
        ax1.bar(bins, temp_df['non_event'], bottom=temp_df['event'], color=colors)
        ax1.set_xticks(bins)
        ax1.set_xticklabels(temp_df['bin'], rotation=45, ha='right')
        fig.subplots_adjust(bottom=0.25)  # Tránh tràn nhãn dưới

        ax2 = ax1.twinx()
        ax2.plot(bins, temp_df['WOE'], color='black', marker='o', linewidth=1.5)
        plt.title(f'WOE Plot for {feature}')
        plt.tight_layout()

        # Lưu biểu đồ vào bộ nhớ và chèn vào Excel
        imgdata = BytesIO()
        plt.savefig(imgdata, format='png')
        imgdata.seek(0)
        plt.close()

        # Chèn hình ảnh vào Excel tại cột L (cột 11), dòng metadata_start_row
        worksheet.insert_image(metadata_start_row, 11, f"{feature}_plot.png", {
            'image_data': imgdata,
            'x_scale': 1.2,
            'y_scale': 1.2
        })

        # Cập nhật current_row: số dòng bảng WOE + khoảng trắng để cách biệt
        current_row += len(woe_df) + 15

    writer.close()  # Đóng file Excel


def transform_to_woe(df, features, target, custom_bins=None, feature=None):
    # Tạo bản sao để tránh thay đổi dữ liệu gốc
    df_transformed = df.copy()

    # Binning biến (nếu cung cấp 1 biến riêng lẻ) để thêm cột GRP_<feature>
    df_transformed = binning_all_variables(df, threshold=0.05, custom_bins=custom_bins, feature=feature)

    # Xác định danh sách biến cần xử lý
    features_to_process = [feature] if feature else features

    # Lặp qua từng biến để tạo biến WOE tương ứng
    for feature in features_to_process:
        # Kiểm tra biến đã được binning chưa (phải có cột GRP_<feature>)
        if 'GRP_' + feature not in df_transformed.columns:
            print(f'Warning: GRP_{feature} không tồn tại trong dữ liệu sau binning!')
            continue

        # Tính bảng WOE và IV từ biến đã binning
        woe_iv, iv = caculate_WOE_IV(df_transformed, 'GRP_' + feature, target)

        # Đảm bảo kiểu dữ liệu bin là chuỗi (tránh lỗi khi map)
        woe_iv['bin'] = woe_iv['bin'].astype(str)
        df_transformed['GRP_' + feature] = df_transformed['GRP_' + feature].astype(str)

        # Tạo mapping từ bin sang WOE
        woe_mapping = woe_iv.set_index('bin')['WOE'].rename('WOE_' + feature)

        # Gán giá trị WOE tương ứng cho từng dòng
        df_transformed['WOE_' + feature] = df_transformed['GRP_' + feature].map(woe_mapping)

        # Đảm bảo kết quả là kiểu float
        df_transformed['WOE_' + feature] = df_transformed['WOE_' + feature].astype(float)

    # Trả về DataFrame đã thêm cột WOE_<feature> tương ứng
    return df_transformed



sl = ['days90_310L', 'sex_738L']
df_transformed = transform_to_woe(df, sl, 'target', custom_bins=None, feature=None)
df_transformed


def get_data_from_sql(server, database, table):
    """
    Kết nối SQL Server và lấy dữ liệu từ bảng chỉ định.

    Args:
        server (str): Tên server SQL.
        database (str): Tên database trong SQL Server.
        table (str): Tên bảng cần lấy dữ liệu.

    Returns:
        pd.DataFrame: Dữ liệu bảng dưới dạng DataFrame.
    """
    # Tạo chuỗi kết nối sử dụng Trusted_Connection (dùng Windows Authentication)
    conn_str = (
        f"DRIVER={{SQL Server}};"
        f"SERVER={server};"
        f"DATABASE={database};"
        f"Trusted_Connection=yes;"
    )

    # Thiết lập kết nối
    conn = pyodbc.connect(conn_str)

    # Câu lệnh truy vấn SQL để lấy toàn bộ bảng
    query = f"SELECT * FROM {table}"

    # Đọc dữ liệu vào DataFrame
    df = pd.read_sql(query, conn)

    # Đóng kết nối
    conn.close()

    # Trả lại DataFrame chứa dữ liệu
    return df



# Khởi tạo từ điển rỗng để lưu IV của từng biến
iv_dict = {}

# Duyệt qua từng cột trong dữ liệu huấn luyện
for feature in df_train_iv.columns:
    # Bỏ qua các cột target và case_id vì không tính IV cho chúng
    if feature in ['target', 'case_id']:
        continue

    try:
        # Tính bảng WOE cho biến hiện tại (đã có hàm woe_feature_table xử lý binning & WOE)
        woe_table = woe_feature_table(df_train_iv, custom_bins=ctb_df, feature=feature)
        
        # Lấy giá trị IV ở dòng 'TOTAL' từ bảng kết quả
        iv = woe_table.loc['TOTAL', 'IV']

        # Ghi lại IV vào từ điển
        iv_dict[feature] = iv

    except Exception as e:
        # Nếu có lỗi xảy ra, in thông báo lỗi kèm tên biến
        print(f"Lỗi khi tính IV cho biến {feature}: {e}")



# Hàm thực hiện lựa chọn biến theo phương pháp forward stepwise greedy

def forward_variable_selection(X_woe, y, iv_dict, 
                                iv_threshold=0.02, corr_threshold=0.5, 
                                verbose=True):
    # Danh sách biến được chọn
    selected_vars = []

    # Lọc các biến còn lại có IV đủ lớn
    remaining_vars = [var for var in X_woe.columns if iv_dict.get(var, 0) > iv_threshold]

    # Lưu lịch sử chọn biến: tên, GINI, số biến đã chọn
    history = []

    # Khởi tạo ma trận X hiện tại rỗng (chỉ index đúng)
    current_X = pd.DataFrame(index=X_woe.index)

    while True:
        best_var = None       # Biến tốt nhất trong vòng hiện tại
        best_gini = -np.inf   # GINI cao nhất
        best_model = None     # Mô hình tốt nhất tạm thời

        for var in remaining_vars:
            # Kiểm tra tương quan với các biến đã chọn
            if selected_vars:
                corrs = X_woe[selected_vars + [var]].corr(method="spearman").iloc[:-1, -1]
                if corrs.abs().max() > corr_threshold:
                    continue  # Bỏ qua nếu tương quan vượt ngưỡng

            # Fit mô hình mới với biến này thêm vào current_X
            temp_X = add_constant(pd.concat([current_X, X_woe[[var]]], axis=1))
            model = Logit(y, temp_X).fit(disp=0)
            pred = model.predict(temp_X)
            auc = roc_auc_score(y, pred)
            gini = 2 * auc - 1

            # Cập nhật nếu GINI tốt hơn
            if gini > best_gini:
                best_gini = gini
                best_var = var
                best_model = model

        # Nếu không còn biến nào được chọn thì dừng
        if best_var is None:
            if verbose:
                print("✅ Dừng: Không còn biến thỏa điều kiện.")
            break

        # Lưu lại biến được chọn tốt nhất
        selected_vars.append(best_var)
        current_X = pd.concat([current_X, X_woe[[best_var]]], axis=1)
        remaining_vars.remove(best_var)

        # Ghi lại lịch sử lựa chọn
        history.append({
            'selected_var': best_var,
            'gini': best_gini,
            'n_vars': len(selected_vars)
        })

        if verbose:
            print(f"✅ Chọn biến: {best_var: <20} | GINI: {best_gini:.4f}")

    # Trả về danh sách biến được chọn và lịch sử theo dạng DataFrame
    return selected_vars, pd.DataFrame(history)


# Hàm áp dụng binning và WOE cho tập test dựa trên thông tin từ tập train

def apply_multiple_bin_woe(df_test, df_train, ctb_dict):
    # Tạo bản sao để không thay đổi dữ liệu gốc
    df_test = df_test.copy()

    # Lấy danh sách các cột GRP_ từ train (đã binning sẵn)
    grp_cols = [col for col in df_train.columns if col.startswith('GRP_')]
    all_varnames = [col.replace('GRP_', '') for col in grp_cols]

    # Chỉ giữ lại những biến gốc có trong test
    varnames = [var for var in all_varnames if var in df_test.columns]

    print(f"[ℹ️] Đang map WOE cho {len(varnames)} biến có trong test: {varnames}")

    # Duyệt từng biến
    for var in varnames:
        try:
            bin_col = f'GRP_{var}'        # Tên cột nhóm
            woe_col = f'WOE_{var}'        # Tên cột WOE

            # Đánh dấu missing
            missing_mask = df_test[var].isna()

            if var in ctb_dict:  # Biến số có bin
                bin_edges = [-np.inf] + ctb_dict[var] + [np.inf]  # Tạo khoảng bin
                df_test[bin_col] = pd.cut(df_test[var], bins=bin_edges, include_lowest=True)
                df_test[bin_col] = df_test[bin_col].astype(object)

                # Gán "Missing" nếu có thiếu
                if missing_mask.any():
                    df_test.loc[missing_mask, bin_col] = 'Missing'

            else:  # Biến phân loại (categorical)
                df_test[bin_col] = df_test[var].astype(object)

                if missing_mask.any():
                    df_test.loc[missing_mask, bin_col] = 'Missing'

            # Trích mapping WOE từ train
            mapping = df_train[[bin_col, woe_col]].drop_duplicates()
            bin_to_woe = {str(k): v for k, v in zip(mapping[bin_col], mapping[woe_col])}

            # Map giá trị WOE tương ứng từ nhóm
            df_test[woe_col] = df_test[bin_col].astype(str).map(bin_to_woe)

        except Exception as e:
            print(f"[⚠️] Lỗi với biến '{var}': {e}")

    # Trả lại dữ liệu test đã gán WOE
    return df_test


# Hàm Psi chia đều số obs bằng nhau
def calculate_psi_percentile(expected, actual, bins=10, verbose=True):
    """
    Tính PSI giữa hai phân phối với cách chia bin theo percentile (mỗi bin có số lượng quan sát gần bằng nhau).

    Args:
        expected: phân phối gốc (vd: train scores)
        actual: phân phối mới (vd: test scores)
        bins: số bin (mặc định 10)
        verbose: nếu True thì in kết quả từng nhóm

    Returns:
        psi_total: tổng PSI
        psi_df: DataFrame chứa PSI từng nhóm
    """
    # 1. Xác định breakpoints theo percentiles từ expected
    quantiles = np.linspace(0, 1, bins + 1)
    breakpoints = np.quantile(expected, quantiles)
    breakpoints[0] = -np.inf  # để bao trùm toàn bộ giá trị
    breakpoints[-1] = np.inf

    # 2. Phân loại bin theo breakpoints
    expected_bins = pd.cut(expected, bins=breakpoints, include_lowest=True)
    actual_bins = pd.cut(actual, bins=breakpoints, include_lowest=True)

    # 3. Tính % từng bin
    expected_percents = expected_bins.value_counts(sort=False, normalize=True).values
    actual_percents = actual_bins.value_counts(sort=False, normalize=True).values

    # 4. Thay epsilon để tránh chia 0
    epsilon = 1e-6
    expected_percents = np.where(expected_percents == 0, epsilon, expected_percents)
    actual_percents = np.where(actual_percents == 0, epsilon, actual_percents)

    # 5. Tính PSI từng bin
    psi_values = (expected_percents - actual_percents) * np.log(expected_percents / actual_percents)
    psi_total = np.sum(psi_values)

    # 6. Tạo bảng PSI
    bin_labels = [f"[{breakpoints[i]:.2f}, {breakpoints[i+1]:.2f})" for i in range(bins)]
    psi_df = pd.DataFrame({
        'bin_range': bin_labels,
        'train_scores_pct': expected_percents,
        'oot_scores_pct': actual_percents,
        'psi_value': psi_values
    })

    if verbose:
        print(psi_df)
        print(f"\n➡ Tổng PSI: {psi_total:.6f}")

    return psi_total, psi_df


# VISUALIZE PSI VÀ XUẤT FILE EXCEL
def plot_psi(psi_detail, psi_total, title="PSI Distribution", excel_file=None):
    """
    Vẽ biểu đồ PSI và (nếu có) xuất ra file Excel.
    psi_detail: DataFrame có các cột ['bin_range', 'train_scores_pct', 'oot_scores_pct', 'psi_value']
    psi_total: tổng PSI
    title: tiêu đề biểu đồ
    excel_file: tên file excel để lưu (vd: 'psi_output.xlsx'), nếu None thì chỉ hiển thị
    """
    # Chuyển tỷ lệ về %
    df = psi_detail.copy()
    df['train_scores_pct'] = df['train_scores_pct'] * 100
    df['oot_scores_pct']   = df['oot_scores_pct'] * 100

    # Vẽ biểu đồ
    fig, ax1 = plt.subplots(figsize=(12,6))
    width = 0.35
    x = range(len(df))

    ax1.bar([i - width/2 for i in x], df['train_scores_pct'], 
            width=width, label='Train %', alpha=0.7)
    ax1.bar([i + width/2 for i in x], df['oot_scores_pct'], 
            width=width, label='OOT %', alpha=0.7)

    ax1.set_xlabel("Bin Range")
    ax1.set_ylabel("Percentage (%)")
    ax1.set_xticks(list(x))
    ax1.set_xticklabels(df['bin_range'], rotation=45, ha='right')
    ax1.legend(loc='upper left')

    ax2 = ax1.twinx()
    ax2.plot(x, df['psi_value'], color='red', marker='o', label='PSI Value')
    ax2.set_ylabel("PSI Value")
    ax2.legend(loc='upper right')

    plt.title(f"{title}\nTotal PSI = {psi_total:.4f}")
    plt.tight_layout()

    # Nếu có excel_file thì lưu biểu đồ vào Excel
    if excel_file:
        img_data = BytesIO()
        fig.savefig(img_data, format='png')
        img_data.seek(0)

        wb = Workbook()
        ws = wb.active
        ws.title = "PSI_Chart"

        img = Image(img_data)
        img.anchor = "A1"
        ws.add_image(img)

        wb.save(excel_file)
        print(f"✅ Biểu đồ đã được lưu vào {excel_file}")

    plt.show()



def calculate_csi_discrete_verbose(expected_series, actual_series, verbose=True):
    """
    Tính CSI cho biến rời rạc + hiển thị chi tiết từng nhóm.

    Returns:
        csi_total: CSI tổng
        csi_df: DataFrame chi tiết từng nhóm
    """
    # Tính phân phối
    expected_dist = expected_series.value_counts(normalize=True)
    actual_dist = actual_series.value_counts(normalize=True)

    expected_dist.index = expected_dist.index.astype(str)
    actual_dist.index = actual_dist.index.astype(str)

    all_bins = sorted(set(expected_dist.index).union(actual_dist.index))
    epsilon = 1e-6

    rows = []
    csi_total = 0

    for bin_val in all_bins:
        p = expected_dist.get(bin_val, 0)
        q = actual_dist.get(bin_val, 0)

        p_adj = max(p, epsilon)
        q_adj = max(q, epsilon)

        csi_bin = (p_adj - q_adj) * np.log(p_adj / q_adj)
        csi_total += csi_bin

        rows.append({
            'bin': bin_val,
            'expected_pct': p,
            'actual_pct': q,
            'csi_value': csi_bin
        })

    csi_df = pd.DataFrame(rows)

    if verbose:
        print(csi_df)
        print(f"\n➡ CSI tổng: {round(csi_total, 4)}")

    return round(csi_total, 4), csi_df



def calculate_csi_for_features_verbose(df_train, df_test, features, verbose=True):
    """
    Tính CSI cho nhiều biến + hiển thị chi tiết từng nhóm.

    Returns:
        csi_summary_df: Bảng tổng CSI
        csi_detail_dict: Dict chứa bảng chi tiết từng biến (đã thêm tên biến)
    """
    csi_results = []
    csi_detail_dict = {}

    for col in features:
        expected_series = df_train[col].astype(str)
        actual_series = df_test[col].astype(str)

        csi_val, csi_df = calculate_csi_discrete_verbose(expected_series, actual_series, verbose=verbose)

        # Thêm cột Feature để dễ nhận biết khi export
        csi_df = csi_df.copy()
        csi_df.insert(0, "Feature", col)

        csi_results.append({'feature': col, 'csi': csi_val})
        csi_detail_dict[col] = csi_df

    csi_summary_df = pd.DataFrame(csi_results).sort_values(by='csi', ascending=False).reset_index(drop=True)
    return csi_summary_df, csi_detail_dict



# VISUALIZE CSI VÀ XUẤT FILE EXCEL
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image

def export_csi_report(csi_summary, csi_details, excel_file="csi_report.xlsx"):
    """
    Xuất CSI detail + chart cho từng biến vào 1 file Excel.
    - Bảng CSI detail từ cột A
    - Biểu đồ CSI từ cột F
    - Mỗi biến cách nhau 10 dòng
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "CSI_Report"

    row = 1  # dòng bắt đầu ghi
    
    for feature, df in csi_details.items():
        # --- Ghi tên biến ---
        ws.cell(row=row, column=1, value=f"Feature: {feature}")
        row += 1

        # --- Ghi bảng CSI detail ---
        # header
        for j, col in enumerate(df.columns, start=1):
            ws.cell(row=row, column=j, value=col)
        row += 1
        # data
        for i in range(len(df)):
            for j, col in enumerate(df.columns, start=1):
                ws.cell(row=row+i, column=j, value=df.iloc[i, j-1])
        row += len(df)

        # --- Xác định tên cột cho train/oot ---
        colnames = [c.lower() for c in df.columns]
        if "train_pct" in colnames:
            train_col, oot_col = "train_pct", "oot_pct"
        elif "train_scores_pct" in colnames:
            train_col, oot_col = "train_scores_pct", "oot_scores_pct"
        elif "expected_pct" in colnames:
            train_col, oot_col = "expected_pct", "actual_pct"
        else:
            raise KeyError(f"Không tìm thấy cột tỷ lệ trong bảng CSI của {feature}: {df.columns.tolist()}")

        # --- Vẽ biểu đồ CSI ---
        fig, ax1 = plt.subplots(figsize=(6,4))
        x = range(len(df))
        width = 0.35

        ax1.bar([i - width/2 for i in x], df[train_col]*100, width=width, alpha=0.7, label="Train %")
        ax1.bar([i + width/2 for i in x], df[oot_col]*100,   width=width, alpha=0.7, label="OOT %")

        ax1.set_ylabel("Percentage (%)")
        ax1.set_xticks(list(x))
        ax1.set_xticklabels(df['bin'], rotation=45, ha='right')
        ax1.legend(loc="upper left")

        if "csi_value" in df.columns:
            ax2 = ax1.twinx()
            ax2.plot(x, df['csi_value'], color="red", marker="o", label="CSI Value")
            ax2.set_ylabel("CSI Value")
            ax2.legend(loc="upper right")

        plt.title(f"CSI - {feature}")
        plt.tight_layout()

        # --- Lưu chart vào bộ nhớ ---
        img_data = BytesIO()
        fig.savefig(img_data, format="png")
        plt.close(fig)
        img_data.seek(0)

        # --- Chèn chart vào Excel ---
        img = Image(img_data)
        img.anchor = f"F{row-len(df)-1}"  # đặt biểu đồ song song với bảng
        ws.add_image(img)

        # --- Tạo khoảng cách 10 dòng ---
        row += 10

    wb.save(excel_file)
    print(f"✅ CSI report đã được lưu tại {excel_file}")

